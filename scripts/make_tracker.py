#!/usr/bin/env python3
"""Generate Daily_Tracker.xlsx with three sheets: Daily, L-Bank, STAR."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUT = "Daily_Tracker.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------- Daily sheet
DAILY_HEADERS = [
    "day", "phase", "topic",
    "theory_done", "hands_on_done",
    "drill_count", "dsa_count", "mock_score", "notes",
]

# Phase + topic per day from appendix/B-30-day-schedule.md
DAILY_ROWS = [
    (1,  "Week 1 — JS/TS + React fundamentals", "JS core: execution context, hoisting, scope, closures, this"),
    (2,  "Week 1 — JS/TS + React fundamentals", "JS async: event loop, Promises, async/await, pLimit/retry"),
    (3,  "Week 1 — JS/TS + React fundamentals", "TypeScript: generics, utility types, discriminated unions, Zod"),
    (4,  "Week 1 — JS/TS + React fundamentals", "React + hooks: render cycle, custom hooks (useDebounce etc.)"),
    (5,  "Week 1 — JS/TS + React fundamentals", "React advanced + perf: reconciliation, memo, Suspense, transitions"),
    (6,  "Week 1 — JS/TS + React fundamentals", "DSA + mock (JS/React, 45 min recorded)"),
    (7,  "Week 1 — Rest", "Rest"),
    (8,  "Week 2 — RN core + new arch + debugging", "RN architecture basics: threading, bridge, Metro, Hermes vs JSC"),
    (9,  "Week 2 — RN core + new arch + debugging", "New Architecture: JSI, Fabric, TurboModules, Codegen, migration"),
    (10, "Week 2 — RN core + new arch + debugging", "Hermes + bundle + cold start: bytecode, inlineRequires, lazy load"),
    (11, "Week 2 — RN core + new arch + debugging", "Performance: FlashList, Reanimated 3, expo-image"),
    (12, "Week 2 — RN core + new arch + debugging", "Native modules: Swift + Kotlin battery module"),
    (13, "Week 2 — RN core + new arch + debugging", "Debugging deep: DevTools, Hermes Inspector, Instruments, ANR"),
    (14, "Week 2 — RN core + new arch + debugging", "DSA + mock (RN deep-dive, 45 min)"),
    (15, "Week 3 — State, nav, offline, test, CI, security", "State management: 3-bucket, RTK, Zustand, React Query"),
    (16, "Week 3 — State, nav, offline, test, CI, security", "Navigation + deep links: typed params, auth gate, Universal Links"),
    (17, "Week 3 — State, nav, offline, test, CI, security", "Offline-first: MMKV, WatermelonDB, outbox, sync, conflicts"),
    (18, "Week 3 — State, nav, offline, test, CI, security", "Testing: Jest, RNTL, Detox, Maestro, native module mocks"),
    (19, "Week 3 — State, nav, offline, test, CI, security", "CI/CD + release: Fastlane, GH Actions, EAS Build/Update"),
    (20, "Week 3 — State, nav, offline, test, CI, security", "Mobile security: Keychain/Keystore, PKCE, pinning, MASVS L2"),
    (21, "Week 3 — State, nav, offline, test, CI, security", "DSA + mixed mock"),
    (22, "Week 4 — System design, behavioral, conversion", "Mobile system design basics: framework, modularization, monorepo"),
    (23, "Week 4 — System design, behavioral, conversion", "Mobile system design fintech: Zerodha, neobank, payments"),
    (24, "Week 4 — System design, behavioral, conversion", "DSA focused: 8 mediums + LRU cache from scratch"),
    (25, "Week 4 — System design, behavioral, conversion", "Behavioral + STAR: 10 stories in 45-sec + 3-min, recorded"),
    (26, "Week 4 — System design, behavioral, conversion", "Resume, LinkedIn, Tier S apply + 10 referral DMs"),
    (27, "Week 4 — System design, behavioral, conversion", "Tier A apply + recruiter screen prep (TMAY, why, CTC)"),
    (28, "Week 4 — System design, behavioral, conversion", "Full mock day: recruiter → DSA → RN → sys design → behavioral"),
    (29, "Week 4 — System design, behavioral, conversion", "Weak-area repair: top 3 gaps from Day 28 + Tier A− apply"),
    (30, "Week 4 — System design, behavioral, conversion", "Cheatsheet review + go live"),
]

# ---------------------------------------------------------------- L-Bank sheet
L_BANK_HEADERS = ["section", "q_num", "fluent_y_n", "last_drilled"]
L_SECTIONS = [
    ("L.2",  "JavaScript core"),
    ("L.3",  "TypeScript"),
    ("L.4",  "React deep dive"),
    ("L.5",  "RN architecture"),
    ("L.6",  "Hermes, Metro, bundle, startup"),
    ("L.7",  "Performance"),
    ("L.8",  "Native modules (Swift + Kotlin)"),
    ("L.9",  "Debugging"),
    ("L.10", "State management"),
    ("L.11", "Navigation + deep linking"),
    ("L.12", "Networking"),
    ("L.13", "Auth, sessions, tokens"),
    ("L.14", "Offline-first + storage"),
    ("L.15", "Push notifications"),
    ("L.16", "Mobile security"),
    ("L.17", "a11y, color, fonts, i18n"),
    ("L.18", "Testing"),
    ("L.19", "CI/CD, EAS, Fastlane, releases"),
    ("L.20", "Observability"),
    ("L.21", "Mobile system design"),
    ("L.22", "DSA"),
    ("L.23", "Behavioral / STAR"),
    ("L.24", "Resume / LinkedIn / negotiation"),
]

# ---------------------------------------------------------------- STAR sheet
STAR_HEADERS = ["story", "45s_recorded", "120s_recorded", "last_rehearsed"]
STAR_STORIES = [
    "1. Largest impact — The 4-second cold start",
    "2. Hardest technical problem — The phantom crash on Android 12",
    "3. Conflict — Native rewrite vs RN at the redesign",
    "4. Failure — The OTA that froze 4% of Android users",
    "5. Leadership without authority — The cross-team auth refactor",
    "6. Mentorship — From new-grad to feature owner in 9 months",
    "7. Ambiguity — Defining 'engagement' for the mobile team",
    "8. Trade-off — Picking React Query over Apollo",
    "9. Saying no — Pushing back on the 'just add WebView' shortcut",
    "10. Customer obsession — The accessibility bug nobody filed",
]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_yn_validation(ws, col_letter, last_row):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")
    ws.add_data_validation(dv)


def add_yn_conditional(ws, col_letter, last_row):
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    rng = f"{col_letter}2:{col_letter}{last_row}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Y"'], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"N"'], fill=red))


def build():
    wb = Workbook()

    # ---- Daily ----
    ws = wb.active
    ws.title = "Daily"
    ws.append(DAILY_HEADERS)
    for day, phase, topic in DAILY_ROWS:
        if day == 1:
            # Sample row showing how to fill a completed day
            ws.append([
                1,
                "Week 1 — JS/TS + React fundamentals",
                "JS core: execution context, hoisting, scope, closures, this",
                "Y", "Y", 10, 5, 7.5,
                "SAMPLE — closures + this puzzles done; stumbled on TDZ + arrow this; redo Q5,Q8 tomorrow.",
            ])
        else:
            ws.append([day, phase, topic, "N", "N", 0, 0, "", ""])
    last = ws.max_row
    style_header(ws, len(DAILY_HEADERS))
    autosize(ws, [6, 42, 60, 13, 14, 12, 11, 12, 50])
    for row in ws.iter_rows(min_row=2, max_row=last, max_col=len(DAILY_HEADERS)):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
    add_yn_validation(ws, "D", last)
    add_yn_validation(ws, "E", last)
    add_yn_conditional(ws, "D", last)
    add_yn_conditional(ws, "E", last)
    # Mock score 0–10 validation
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_score = DataValidation(type="decimal", operator="between", formula1=0, formula2=10, allow_blank=True)
    dv_score.add(f"H2:H{last}")
    ws.add_data_validation(dv_score)
    # Summary row
    summary_row = last + 2
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=f'=COUNTIF(D2:D{last},"Y")&"/"&{last-1}')
    ws.cell(row=summary_row, column=5, value=f'=COUNTIF(E2:E{last},"Y")&"/"&{last-1}')
    ws.cell(row=summary_row, column=6, value=f"=SUM(F2:F{last})")
    ws.cell(row=summary_row, column=7, value=f"=SUM(G2:G{last})")
    ws.cell(row=summary_row, column=8, value=f'=IFERROR(AVERAGEIF(H2:H{last},">0"),"")')
    for c in range(1, 9):
        ws.cell(row=summary_row, column=c).font = Font(bold=True)

    # ---- L-Bank ----
    ws2 = wb.create_sheet("L-Bank")
    ws2.append(L_BANK_HEADERS)
    for code, name in L_SECTIONS:
        section_label = f"{code} — {name}"
        for q in range(1, 51):
            if code == "L.2" and q == 1:
                # Sample row
                ws2.append([section_label, q, "Y", "2026-05-10"])
            else:
                ws2.append([section_label, q, "N", ""])
    last2 = ws2.max_row
    style_header(ws2, len(L_BANK_HEADERS))
    autosize(ws2, [44, 8, 12, 16])
    add_yn_validation(ws2, "C", last2)
    add_yn_conditional(ws2, "C", last2)
    ws2.auto_filter.ref = f"A1:D{last2}"

    # L-Bank summary sheet
    ws_sum = wb.create_sheet("L-Bank Summary")
    ws_sum.append(["section", "fluent", "total", "% fluent"])
    for i, (code, name) in enumerate(L_SECTIONS, start=2):
        section_label = f"{code} — {name}"
        ws_sum.cell(row=i, column=1, value=section_label)
        ws_sum.cell(row=i, column=2,
                    value=f'=COUNTIFS(\'L-Bank\'!A:A,A{i},\'L-Bank\'!C:C,"Y")')
        ws_sum.cell(row=i, column=3,
                    value=f"=COUNTIF('L-Bank'!A:A,A{i})")
        ws_sum.cell(row=i, column=4, value=f"=IFERROR(B{i}/C{i},0)")
        ws_sum.cell(row=i, column=4).number_format = "0%"
    total_row = len(L_SECTIONS) + 2
    ws_sum.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_sum.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})")
    ws_sum.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})")
    ws_sum.cell(row=total_row, column=4, value=f"=IFERROR(B{total_row}/C{total_row},0)")
    ws_sum.cell(row=total_row, column=4).number_format = "0%"
    for c in range(1, 5):
        ws_sum.cell(row=total_row, column=c).font = Font(bold=True)
    style_header(ws_sum, 4)
    autosize(ws_sum, [44, 10, 10, 12])

    # ---- STAR ----
    ws3 = wb.create_sheet("STAR")
    ws3.append(STAR_HEADERS)
    for i, s in enumerate(STAR_STORIES):
        if i == 0:
            # Sample row
            ws3.append([s, "Y", "Y", "2026-05-10"])
        else:
            ws3.append([s, "N", "N", ""])
    last3 = ws3.max_row
    style_header(ws3, len(STAR_HEADERS))
    autosize(ws3, [60, 16, 16, 18])
    add_yn_validation(ws3, "B", last3)
    add_yn_validation(ws3, "C", last3)
    add_yn_conditional(ws3, "B", last3)
    add_yn_conditional(ws3, "C", last3)
    for row in ws3.iter_rows(min_row=2, max_row=last3, max_col=len(STAR_HEADERS)):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER

    # ---- README ----
    ws_r = wb.create_sheet("README", 0)
    notes = [
        ["React Native Interview Prep — Daily Tracker"],
        [""],
        ["Sheets:"],
        ["  • Daily         — 30-day plan; mark theory_done / hands_on_done = Y/N, log drills, DSA, mock score, notes."],
        ["  • L-Bank        — 1,150 drill questions (23 sections × 50). Mark fluent_y_n = Y once you can answer in ≤60s."],
        ["  • L-Bank Summary— Auto-rolled fluency % per section."],
        ["  • STAR          — 10 stories from Appendix E. Mark 45s and 120s recordings + last rehearsed date."],
        [""],
        ["Conventions:"],
        ["  • Y/N cells use data validation + green/red conditional formatting."],
        ["  • Date cells (last_drilled / last_rehearsed): enter as YYYY-MM-DD."],
        ["  • mock_score: 0–10 (validated)."],
        [""],
        ["Source: appendix/B-30-day-schedule.md, appendix/L-50q-drill-bank.md, appendix/E-star-story-bank.md"],
    ]
    for row in notes:
        ws_r.append(row)
    ws_r["A1"].font = Font(bold=True, size=14)
    ws_r.column_dimensions["A"].width = 110

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    build()
